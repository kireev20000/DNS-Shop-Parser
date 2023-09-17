import pickle
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from tqdm import tqdm
from random import randint
from datetime import datetime
from time import sleep as pause
from bs4 import BeautifulSoup
import undetected_chromedriver as uc


def parse_characteristics_page(driver, url):
    """ Парсит страницу товара по ссылке."""
    driver.get(url)
    pause(randint(7, 11))
    soup = BeautifulSoup(driver.page_source, 'lxml')
    #print(soup.prettify())

    name = soup.find('div', class_="product-card-description__title")
    price = soup.find('div', class_="product-buy__price")
    desc = soup.find('div', class_="product-card-description-text")
    avail = soup.find('a', class_="order-avail-wrap__link ui-link ui-link_blue")
    charcs = soup.find_all('div', class_="product-characteristics__spec-title")
    cvalue = soup.find_all('div', class_="product-characteristics__spec-value")
    main_picture = soup.find('img', class_="product-images-slider__main-img")
    pictures_soup = soup.find_all('img', class_="product-images-slider__img loaded tns-complete")

    pictures_list = []
    for i in pictures_soup:
        _ = pictures_list.append(i.get('data-src'))
        if _ is not None:
            pictures_list.append(_)

    span_tags = soup.find_all('span')
    for i in span_tags:
        if bool(str(i).find('data-go-back-catalog') != -1):
            category = i

    tech_spec = {}
    for f1, f2 in zip(charcs, cvalue):
        tech_spec[f1.text.rstrip().lstrip()] = f2.text.rstrip().lstrip()

    notebook = {}

    notebook["Категория"] = category.text.lstrip(': ')
    notebook["Наименование"] = name.text[15:]
    notebook["Цена"] = int(price.text.replace(' ', '')[:-1])
    notebook["Доступность"] = avail.text if avail is not None else 'Товара нет в наличии'
    notebook["Ссылка на товар"] = url
    notebook["Описание"] = desc.text
    notebook["Главное изображение"] = main_picture.get('src') if main_picture is not None else 'У товара нет картинок'
    notebook["Лист с картинками"] = pictures_list
    notebook["Характеристики"] = list(tech_spec.items())

    # for i, j in notebook.items():
    #     print(i, j)
    return notebook


def get_all_category_page_urls(driver, url_to_parse):
    """ Получаем URL категории и парсим ссылки с неё."""
    page = 1
    url = url_to_parse.format(page=page)
    driver.get(url=url)
    pause(10)

    soup = BeautifulSoup(driver.page_source, 'lxml')

    span_tags = soup.find_all('span')
    for i in span_tags:
        if bool(str(i).find('data-role="items-count"') != -1):
            number_of_pages = [int(x) for x in str(i) if x.isdigit()]

    res = int(''.join(map(str, number_of_pages)))
    pages_total = ((res // 18) + 1)
    print(f'Всего в категории {pages_total+1} страницы')

    urls = []

    while True:
        page_urls = get_urls_from_page(driver)
        urls += page_urls

        if page >= pages_total:
            break

        page += 1
        url = url_to_parse.format(page=page)
        driver.get(url)
        pause(randint(6, 9))

    return urls


def get_urls_from_page(driver):
    """ Собирает все ссылки на текущей странице. """
    soup = BeautifulSoup(driver.page_source, 'lxml')
    elements = soup.find_all('a', class_="catalog-product__name ui-link ui-link_black")
    return list(map(
        lambda element: 'https://www.dns-shop.ru' + element.get("href") + 'characteristics/',
        elements
    ))


def to_excel(data, file_name="table"):

    workbook = Workbook()
    sheet = workbook.active
    print('=' * 20)
    print('Начался экспорт в Excel Таблицу')

    column_names = [
            "Категория",
            "Наименование",
            "Цена",
            "Доступность",
            "Ссылка на товар",
            "Описание",
            "Главное изображение",
            "Лист с картинками",
            "Характеристики",
    ]

    side = Side(border_style='thin')
    border = Border(
        left=side,
        right=side,
        top=side,
        bottom=side
    )
    alignment = Alignment(
        horizontal='center',
        vertical='center'
    )

    for column, name in enumerate(column_names, 1):
        cell = sheet.cell(
            column=column,
            row=1,
            value=name
        )
        cell.font = Font(bold=True)
        cell.border = border
        cell.alignment = alignment

    counter = 1
    for index, value in enumerate(data, 2):
        for i in value.values():
            cell = sheet.cell(
                column=counter,
                row=index,
                value=str(i) if type(i) == list else i
            )
            cell.alignment = Alignment(horizontal='left')
            counter += 1

        counter = 1

    for i in 'ABCDEFGHI':
        sheet.column_dimensions[i].width = 30

    workbook.save(f"{file_name} {datetime.now().strftime('%d.%m.%y %H-%M-%S')}.xlsx")


def main():

    driver = uc.Chrome()
    urls_to_parse = [
        'https://www.dns-shop.ru/catalog/recipe/e585499db2f27251/demontaz/?p={page}',
        'https://www.dns-shop.ru/catalog/17a89bb916404e77/platy-rasshireniya/?p={page}',
        'https://www.dns-shop.ru/catalog/c8a984d0ba7f4e77/radiosistemy/?p={page}',
        'https://www.dns-shop.ru/catalog/2c0f47131ade2231/aksessuary-dlya-materinskix-plat/?p={page}',
        'https://www.dns-shop.ru/catalog/17a89b8416404e77/karty-videozaxvata/?p={page}',
    ]

    urls = []
    for index, url in enumerate(urls_to_parse):
        print(f'Получение списка всех ссылок из {index+1} категории:')
        parsed_url = get_all_category_page_urls(driver, url)
        urls.append(parsed_url)

    print("Запись всех ссылок в файл url.txt:")
    with open('urls.txt', 'w') as file:
        for url in urls:
            for link in url:
                file.write(link + "\n")

    with open('urls.txt', 'r') as file:
        urls = list(map(lambda line: line.strip(), file.readlines()))
        print(urls)
        info_dump = []
        for url in tqdm(urls, ncols=70, unit='товаров',
                        colour='blue', file=sys.stdout):
            info_dump.append(parse_characteristics_page(driver, url))

    with open('dump_list_pickle.txt', 'wb+') as file:
        pickle.dump(info_dump, file)

    with open('dump_list_pickle.txt', 'rb') as file:
        info_dump = pickle.load(file)

    to_excel(info_dump, file_name="info_dump")


if __name__ == '__main__':
    main()
    print('=' * 20)
    print('Все готово!')
